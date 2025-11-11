import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from typing import Optional

from app.aggregate import DataAggregator


class ExcelExporter:
    def __init__(self, aggregator: DataAggregator):
        self.aggregator = aggregator
        self.output_dir = os.getenv('EXPORT_DIR', './data/processed')
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_excel_workbook(self, output_filename: Optional[str] = None) -> str:
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"hackathon_aggregations_{timestamp}.xlsx"
        
        output_path = os.path.join(self.output_dir, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            top_techs = self.aggregator.get_top_technologies(limit=50)
            if not top_techs.empty:
                top_techs.to_excel(writer, sheet_name='Top Technologies', index=False)
            
            top_skills = self.aggregator.get_top_skills(limit=50)
            if not top_skills.empty:
                top_skills.to_excel(writer, sheet_name='Top Skills', index=False)
            
            submissions_by_hackathon = self.aggregator.get_submissions_by_hackathon()
            if not submissions_by_hackathon.empty:
                submissions_by_hackathon.to_excel(writer, sheet_name='Submissions by Hackathon', index=False)
            
            team_size = self.aggregator.get_team_size_distribution()
            if not team_size.empty:
                team_size.to_excel(writer, sheet_name='Team Size Distribution', index=False)
            
            countries = self.aggregator.get_country_distribution(limit=50)
            if not countries.empty:
                countries.to_excel(writer, sheet_name='Country Distribution', index=False)
            
            occupations = self.aggregator.get_occupation_breakdown(limit=50)
            if not occupations.empty:
                occupations.to_excel(writer, sheet_name='Occupation Breakdown', index=False)
            
            specialty = self.aggregator.get_specialty_distribution()
            if not specialty.empty:
                specialty.to_excel(writer, sheet_name='Specialty Distribution', index=False)
            
            work_exp = self.aggregator.get_work_experience_distribution()
            if not work_exp.empty:
                work_exp.to_excel(writer, sheet_name='Work Experience', index=False)
            
            time_trends = self.aggregator.get_time_trends(period='daily')
            if not time_trends.empty:
                time_trends.to_excel(writer, sheet_name='Time Trends', index=False)
            
            summary_stats = self.aggregator.get_summary_statistics()
            summary_df = pd.DataFrame([
                {'Metric': 'Total Submissions', 'Value': summary_stats['total_submissions']},
                {'Metric': 'Total Registrants', 'Value': summary_stats['total_registrants']},
                {'Metric': 'Unique Hackathons', 'Value': summary_stats['unique_hackathons']},
                {'Metric': 'Unique Organizations', 'Value': summary_stats['unique_organizations']},
                {'Metric': 'Date Range Start', 'Value': summary_stats['date_range']['start']},
                {'Metric': 'Date Range End', 'Value': summary_stats['date_range']['end']},
                {'Metric': 'Most Popular Technology', 'Value': summary_stats['most_popular_technology']},
                {'Metric': 'Most Popular Skill', 'Value': summary_stats['most_popular_skill']},
                {'Metric': 'Top Country', 'Value': summary_stats['top_country']},
                {'Metric': 'Average Team Size', 'Value': summary_stats['avg_team_size']}
            ])
            summary_df.to_excel(writer, sheet_name='Summary Statistics', index=False)
        
        self._format_workbook(output_path)
        
        return output_path
    
    def _format_workbook(self, file_path: str) -> None:
        wb = load_workbook(file_path)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            ws.freeze_panes = 'A2'
        
        wb.save(file_path)
    
    def get_export_history(self) -> pd.DataFrame:
        if not os.path.exists(self.output_dir):
            return pd.DataFrame(columns=['Filename', 'Created At', 'Size (KB)'])
        
        files = []
        for filename in os.listdir(self.output_dir):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(self.output_dir, filename)
                file_stat = os.stat(file_path)
                files.append({
                    'Filename': filename,
                    'Created At': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'Size (KB)': round(file_stat.st_size / 1024, 2)
                })
        
        df = pd.DataFrame(files)
        if not df.empty:
            df = df.sort_values('Created At', ascending=False)
        
        return df
    
    def delete_export(self, filename: str) -> bool:
        file_path = os.path.join(self.output_dir, filename)
        
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                return True
            except Exception as e:
                return False
        
        return False
